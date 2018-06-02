# -*- coding: utf-8 -*-
"""
Created on Fri Dec 29 17:26:45 2017

@author: Laura.Jones
"""
from openpyxl import load_workbook
from Bio import SeqIO
import re
import time
from collections import Counter

#for importing the taxon info from excel spreadsheets, telling it the relevant sheet and column to look in.
#cleans up the input of anything beyond 'Genus species'
#doesn't pull info if only 'Genus' with no species
def loadspreadsheet(filename, sheetname, columnno):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    species = []
    for row in sheet.iter_rows(row_offset=1):
        if row[columnno].value != None:
            spp = row[columnno].value.split("'")[0].split('[')[0].split('(')[0].lstrip('?x').rstrip('=').strip()
            if len(spp.split())>1:
                species.append(spp)
    print(filename, 'imported')
    print('Number of species:',len(species))
    species = list(set(species))
    species.sort()
    print('Number of unique species:', len(species))
    return species

#for importing family and genus info from a spreadsheet
def loadfam(filename, sheetname):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    genus = {}
    for row in sheet.iter_rows(row_offset=1):
        if row[1].value != None:
            spp = row[1].value.split("'")[0].split('[')[0].split('(')[0].lstrip('?x').rstrip('=').strip()
            family = row[0].value.strip()
            gen = spp.split()[0]
            if not gen in genus:
                genus[gen] = family
    print(filename, 'imported')
    return genus

#for creating a file output of the species list
def writelist(filename, writelist):
    fd = open(filename, 'w')
    for spp in writelist:
        fd.write('%s\n' % spp)
    fd.close()
    
#for creating a list of the unique genera contained within a species list
def genuslist(specieslist):
    glist = []
    for spp in specieslist:
        glist.append(spp.split()[0])
    return list(set(glist))

#for pulling out plant family info from the 'taxonomy' info in the Genbank file format
def find_family(taxons):
    regex=re.compile(".*(ceae).*")
    found = [m.group(0) for l in taxons for m in [regex.search(l)] if m]
    if len(found) == 0:
        print(taxons)
        return 'unknown'
    return found[0]

#load the taxa information from excel spreadsheets into lists
bcukspecies = loadspreadsheet('barcode-uk-full-species-list.xlsx', 'All species', 5)
irisspecies = loadspreadsheet('birisinfo.xlsx', 'Data', 1)
bcalien = loadspreadsheet('Barcode-Alien.xlsx', 'By species', 2)
flowersurvey = loadspreadsheet('bflowersurveynosp.xlsx', 'DATA', 8)

#concatenate the species lists, keeping the unique entries
fullspecies = list(set(bcukspecies+irisspecies+bcalien+flowersurvey))
print('Number of unique species from all lists:', len(fullspecies))
#write the concatenated list to a .csv
writelist('fullspecieslist.csv', fullspecies)

start_time = time.time()


#Using a genbank type file with all rbcL records from Genbank downloaded:
#keep records that match with species in the full species list
#and write to a new fasta file, with this info: ">Accession|Genus species|Family

keepgb = 0
gborganism = []
gbaccessions = []

with open("genbank-fasta-selected.fas", "w") as fd:
    for index, record in enumerate(SeqIO.parse('ncbi_rbcl_plants_genbank.gb', 'genbank')):
        organism = record.annotations['organism']
        accession = record.annotations['accessions']
        gborganism.append(organism)
        for spp in fullspecies:
            if spp in organism:
                if not accession in gbaccessions:
                    fd.write('>%s|%s|%s\n' % (record.annotations['accessions'][0],organism,find_family(record.annotations['taxonomy'])))
                    fd.write('%s\n' % record.seq)
                    keepgb += 1
                    gbaccessions.append(accession)
    
print(keepgb)

took = time.time() - start_time
print((took)/60, 'minutes')


#create a list of the records kept from the species level pass through the Genbank rbcl database
descriptionlist = []

for index, record in enumerate(SeqIO.parse('genbank-fasta-selected.fas', 'fasta')):
    descriptionlist.append(record.description)

writelist('descriptionlist.csv', descriptionlist)

#compare between the fullspecies and the sequence records that were selected
#working out how many of the records pulled are unique species

descriptionlistspp = []
for spp in descriptionlist:
    descriptionlistspp.append(spp.split('|')[1].strip())

print('Number of NCBI records pulled:', len(descriptionlistspp))
uniquedescriptionlistspp = list(set(descriptionlistspp))
print('Number of unique NCBI records pulled:', len(uniquedescriptionlistspp))

uniquedescriptionlistgenus = genuslist(uniquedescriptionlistspp)

recordcheck = []
for spp in fullspecies:
    if spp in descriptionlistspp:
        recordcheck.append(spp)

print('Number of species from full species list represented in NCBI:', len(recordcheck))
print('Species level coverage:''{percent:.2%}'.format(percent=(len(recordcheck)/len(fullspecies))))

#create a list of the species from the full species list that are not represented in NCBI
noseqrecords = []
for spp in fullspecies:
    if not spp in descriptionlistspp:
        noseqrecords.append(spp)

noseqrecords.sort()
writelist('noseqrecords-for-spp-in-NCBI.csv', noseqrecords)

#create a list of the unique genus contained in the total unique species list 
fullgenus = genuslist(fullspecies)

#creating a list of the genera that are not represented at species level and sort
nogenus = []
for genus in fullgenus:
    if not genus in uniquedescriptionlistgenus:
        nogenus.append(genus)

nogenus.sort()
writelist('noseqrecords-for-genus-in-NCBI-at-spplevel.csv', nogenus)

descriptionlistaccessions = []
for acc in descriptionlist:
    descriptionlistaccessions.append(acc.split('|')[0].strip())


#create a fasta for the genera not represented at species level
keepgengb = 0
gborganismgenus = []
genaccessions = []

with open("genbank-fasta-selected-genus.fas", "w") as fd:
    for index, record in enumerate(SeqIO.parse('ncbi_rbcl_plants_genbank.gb', 'genbank')):
        organism = record.annotations['organism']
        accession = record.annotations['accessions']
        gborganismgenus.append(organism)
        for gen in nogenus:
            if re.search(r'\b%s\b' % (gen), organism):
                if not accession in genaccessions:
                    if accession not in descriptionlistaccessions:
                        fd.write('>%s|%s|%s\n' % (record.annotations['accessions'][0],record.annotations['organism'],find_family(record.annotations['taxonomy'])))
                        fd.write('%s\n' % record.seq)
                        genaccessions.append(accession)
                        keepgengb += 1
    
print(keepgengb)


genusdesclist = []
for index, record in enumerate(SeqIO.parse('genbank-fasta-selected-genus.fas', 'fasta')):
    genusdesclist.append(record.description)
            
genusdesclistspp = []
for gen in genusdesclist:
    genusdesclistspp.append(gen.split('|')[1].strip())
    
genusdesclistgen = genuslist(genusdesclistspp)

writelist('genusremaining.csv', genusdesclistgen)

#create a list of the genera that are not represented in NCBI
nogenseqrecords = []
for gen in nogenus:
    if not gen in genusdesclistgen:
        nogenseqrecords.append(gen)
        
writelist('noseqrecords-for-genera-in-NCBI.csv', nogenseqrecords)

print('Genus coverage:','{percent:.2%}'.format(percent=(len(genusdesclistgen+uniquedescriptionlistgenus)/len(fullgenus))))

#Finding out which families aren't represented in the pulled NCBI records

familylist = []

for index, record in enumerate(SeqIO.parse('genbank-fasta-selected.fas', 'fasta')):
    familylist.append(record.description.split('|')[2].strip())
    

#create a count of the different plant families represented in the database and save

familycount = Counter(familylist)

with open('familycount.csv', 'w') as fd:
    for key in familycount.keys():
        fd.write("%s,%d\n" % (key, familycount[key]))


#find out of the genera that are not available in genbank, which aren't represented at family level

familyref = loadfam('familyspeciesinfo.xlsx', 'Sheet1')

#make a list of all the unique families represented in the pulled records from the species level selection and the genus level

familylistspp = []

for index, record in enumerate(SeqIO.parse('genbank-fasta-selected.fas', 'fasta')):
    familylistspp.append(record.description.split('|')[2].strip())

uniquefamilylistspp = list(set(familylistspp))
print(len(uniquefamilylistspp))

familylistgen = []

for index, record in enumerate(SeqIO.parse('genbank-fasta-selected-genus.fas', 'fasta')):
    familylistgen.append(record.description.split('|')[2].strip())

uniquefamilylistgen = list(set(familylistgen))
print(len(uniquefamilylistgen))

totalfamilylist = list(set(uniquefamilylistspp+uniquefamilylistgen))
print(len(totalfamilylist))

familyofnogenseq = []

for gen in nogenseqrecords:
    if gen in familyref:
        familyofnogenseq.append(familyref[gen])
    else:
        print('no genus', gen)
        
actuallynoseqfamily = []
for family in familyofnogenseq:
    if not family in totalfamilylist:
        actuallynoseqfamily.append(family)
        
writelist('no-family-seq-record.csv', actuallynoseqfamily)

#add together the two fasta files

total1 = 0
total2 = 0

with open("restrictedreferencelibrary.fas", "w") as fd:
    for index, record in enumerate(SeqIO.parse('genbank-fasta-selected-genus.fas', 'fasta')):
         fd.write('>%s\n' % record.description)
         fd.write('%s\n' % record.seq)
         total1 += 1
    for index2, record in enumerate(SeqIO.parse('genbank-fasta-selected.fas', 'fasta')):
         fd.write('>%s\n' % record.description)
         fd.write('%s\n' % record.seq)
         total2 += 1

print(total1)
print(total2)






    









